import re
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from workouts.models import Exercise, MuscleGroup

XLSX_PATH = Path(settings.BASE_DIR).parent / "workout_sample.xlsx"

# Plain-language cleanup for workout_sample.xlsx's Major/Minor Muscle Group
# columns: strip clinical parentheticals ("Chest (Pectoralis Major)" ->
# "Chest") and normalize a few inconsistent/overly-clinical tokens the sheet
# uses interchangeably (singular vs plural deltoid names, "Erector Spinae",
# ...) to the plain term the spec asks for.
SKIP_TOKENS = {"Cardio"}
RENAME_TOKENS = {
    "Front Deltoid": "Front Deltoids",
    "Rear Deltoid": "Rear Deltoids",
    "Side Deltoid": "Side Deltoids",
    "Erector Spinae": "Lower Back",
    "Rectus Abdominis": "Core",
    "Transverse Abdominis": "Core",
    "Lower Abs": "Core",
    "Forearms/Grip": "Forearms",
    "Full Body Stability": "Core",
}


def clean_muscle_list(raw):
    if not raw or raw.strip() in ("", "-"):
        return []
    # Strip parentheticals from the whole string first - a handful of entries
    # (e.g. "Legs (Quadriceps, Hamstrings)") have a comma *inside* the
    # parenthetical, which would otherwise get mis-split into two garbage
    # tokens ("Legs (Quadriceps", "Hamstrings)") if split before stripping.
    raw = re.sub(r"\s*\([^)]*\)", "", raw)
    tokens = []
    for part in raw.split(","):
        part = part.strip()
        part = RENAME_TOKENS.get(part, part)
        if part and part not in SKIP_TOKENS:
            tokens.append(part)
    return tokens


# Not present in workout_sample.xlsx - assigned by hand, per exercise, since a
# generic keyword heuristic misclassifies too many edge cases (e.g. equipment
# alone can't tell a Pull-Up from a Lat Pulldown). Anything absent from this
# map defaults to Intermediate.
ADVANCED = {
    "Barbell Back Squat", "Front Squat", "Deadlift", "Sumo Deadlift", "Pull-Up",
    "Chin-Up", "Bulgarian Split Squat", "Clean and Jerk", "Snatch", "Push Press",
    "Hanging Leg Raise", "Ab Wheel Rollout", "Turkish Get-Up",
}
BEGINNER = {
    "Dumbbell Bench Press", "Dumbbell Flyes", "Push-Up", "Pec Deck Machine",
    "Smith Machine Bench Press", "Lat Pulldown", "Seated Cable Row",
    "Single-Arm Dumbbell Row", "Face Pull", "Straight-Arm Pulldown",
    "Back Extension (Hyperextension)", "Leg Press", "Leg Extension",
    "Leg Curl (Lying/Seated)", "Glute Bridge", "Standing Calf Raise",
    "Seated Calf Raise", "Goblet Squat", "Hack Squat (Machine)", "Step-Up",
    "Hip Abduction Machine", "Hip Adduction Machine", "Dumbbell Shoulder Press",
    "Lateral Raise", "Front Raise", "Rear Delt Fly", "Barbell/Dumbbell Shrug",
    "Cable Lateral Raise", "Reverse Fly (Machine)", "Barbell Curl", "Dumbbell Curl",
    "Hammer Curl", "Preacher Curl", "Concentration Curl", "Cable Curl",
    "Incline Dumbbell Curl", "EZ-Bar Curl", "Triceps Pushdown",
    "Skull Crusher (Lying Triceps Extension)", "Overhead Triceps Extension",
    "Triceps Dip (Bench)", "Cable Overhead Triceps Extension", "Wrist Curl",
    "Reverse Wrist Curl", "Crunch", "Sit-Up", "Plank", "Russian Twist",
    "Bicycle Crunch", "Side Plank", "Cable Crunch", "Mountain Climbers",
    "Battle Ropes", "Treadmill Running", "Stationary Bike", "Rowing Machine",
    "Elliptical Trainer", "Stair Climber",
}

ALTERNATIVES_PER_EXERCISE = 4


class Command(BaseCommand):
    help = "Seed the Exercise reference bank from workout_sample.xlsx (~100 exercises)."

    def handle(self, *args, **options):
        if not XLSX_PATH.exists():
            raise CommandError(f"{XLSX_PATH} not found")

        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        exercises_by_name = {}
        categories_by_name = {}
        created_count = 0
        updated_count = 0

        for _num, name, category, equipment, major, minor in rows:
            if not name:
                continue
            name = str(name).strip()
            difficulty = (
                Exercise.Difficulty.ADVANCED
                if name in ADVANCED
                else Exercise.Difficulty.BEGINNER
                if name in BEGINNER
                else Exercise.Difficulty.INTERMEDIATE
            )

            exercise, created = Exercise.objects.update_or_create(
                name=name,
                defaults={
                    "equipment": (equipment or "").strip(),
                    "difficulty_level": difficulty,
                },
            )
            primary = [MuscleGroup.objects.get_or_create(name=n)[0] for n in clean_muscle_list(major)]
            secondary = [MuscleGroup.objects.get_or_create(name=n)[0] for n in clean_muscle_list(minor)]
            exercise.primary_muscle_groups.set(primary)
            exercise.secondary_muscle_groups.set(secondary)

            exercises_by_name[name] = exercise
            categories_by_name[name] = (category or "").strip()
            created_count += created
            updated_count += not created

        # Alternatives: other exercises in the same category, ranked by how many
        # primary muscle groups they share, capped at ALTERNATIVES_PER_EXERCISE.
        for name, exercise in exercises_by_name.items():
            own_primary = set(exercise.primary_muscle_groups.values_list("id", flat=True))
            candidates = [
                (other_name, other)
                for other_name, other in exercises_by_name.items()
                if other_name != name and categories_by_name[other_name] == categories_by_name[name]
            ]
            candidates.sort(
                key=lambda pair: len(
                    own_primary & set(pair[1].primary_muscle_groups.values_list("id", flat=True))
                ),
                reverse=True,
            )
            exercise.alternatives.set([other for _, other in candidates[:ALTERNATIVES_PER_EXERCISE]])

        # Drop any previously-seeded exercise no longer in the sheet (e.g. an
        # earlier hand-picked name superseded by this sheet's own naming), so
        # reruns don't accumulate stale near-duplicates.
        stale = Exercise.objects.exclude(name__in=exercises_by_name.keys())
        stale_count = stale.count()
        stale.delete()

        self.stdout.write(
            self.style.SUCCESS(
                f"Seeded Exercises: {created_count} created, {updated_count} updated, {stale_count} stale removed."
            )
        )
